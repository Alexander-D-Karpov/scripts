import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="pymorphy2")

import re
import sys
from natasha import (
    Segmenter, MorphVocab, NewsEmbedding, NewsNERTagger,
    NamesExtractor, Doc
)


class RuPIIMasker:
    def __init__(self):
        self.segmenter = Segmenter()
        self.morph_vocab = MorphVocab()
        self.emb = NewsEmbedding()
        self.ner_tagger = NewsNERTagger(self.emb)
        self.names_extractor = NamesExtractor(self.morph_vocab)

        self.regex_patterns = [
            ('EMAIL', re.compile(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b')),
            ('PHONE', re.compile(
                r'(?:\+7|8)[\s\-]?\(?\d{3,4}\)?[\s\-]?\d{2,3}[\s\-]?\d{2}[\s\-]?\d{2}(?:\s*,?\s*доб\.?\s*\d+(?:-\d+)?)?')),
            ('PHONE', re.compile(r'(?:тел\.?|т\.)\s*8[-\s]?\d{3}[-\s]?\d{3}[-\s]?\d{2}[-\s]?\d{2}')),
            ('SNILS', re.compile(r'\b\d{3}[-\s]?\d{3}[-\s]?\d{3}[-\s]?\d{2}\b')),
            ('INN', re.compile(r'(?i)\b(?:инн\s*[:\-]?\s*)(\d{10}|\d{12})\b')),
            ('PASSPORT', re.compile(r'(?i)(?:паспорт|серия)\s*[:\-]?\s*\d{2}\s?\d{2}\s*(?:№|номер)?\s*\d{6}')),
            ('DATE', re.compile(r'\b\d{1,2}[./]\d{1,2}[./]\d{2,4}\s*г?\.?\b')),
            ('DOC_NUM', re.compile(r'(?:№|N|#)\s*[A-Za-zА-Яа-яЁё]{0,5}[-\s]?\d{3,}')),
            ('DOC_REF', re.compile(r'(?i)(?:в/х|исх|вх|рег)\s*[.№#]\s*\d+')),
            ('CONTRACT_12', re.compile(r'\b\d{12}\b')),
            ('FIO', re.compile(
                r'(?i)(?:уважаем(?:ый|ая|ое|ые)|дорог(?:ой|ая|ое|ие))\s+([А-ЯЁ][а-яё]+(?:\s+[А-ЯЁ][а-яё]+){0,2})(?=[!\s,.])')),
            ('FIO', re.compile(r'(?i)\bот\s+([А-ЯЁ][а-яё]{2,}\s+[А-ЯЁ]\.?\s*[А-ЯЁ]?\.?)(?=\s|$|[,.])')),
            ('FIO', re.compile(
                r'(?i)(?:начальник[аеуом]*|директор[ауе]*|менеджер[ауе]*|специалист[ауе]*|инженер[ауе]*|мастер[ауе]*)[^–\-]{0,50}[–\-]\s*([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+(?:\s+[А-ЯЁ][а-яё]+)?)')),
            ('FIO', re.compile(r'\b([А-ЯЁ][а-яё]{2,})\s+([А-ЯЁ])\.?\s*([А-ЯЁ])\.?(?=\s|$|[,.\"\'])')),
            ('FIO', re.compile(r'\b([А-ЯЁ])\.?\s*([А-ЯЁ])\.?\s+([А-ЯЁ][а-яё]{2,})(?=\s|$|[,.\"\'])')),
            ('FIO', re.compile(r'(?i)\bя\s*,\s*([А-ЯЁ][а-яё]{2,}\s+[А-ЯЁ][а-яё]{2,}\s+[А-ЯЁ][а-яё]{2,})\s*,')),
            ('FIO', re.compile(r'(?<=[–\-]\s)([А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+)(?=\s*[,.]|\s*$)')),
            ('FIO_FULL', re.compile(
                r'(?<![А-Яа-я])([А-ЯЁ][а-яё]{2,})\s+([А-ЯЁ][а-яё]{2,})\s+([А-ЯЁ][а-яё]{2,}(?:вич|вна|ич|ична))(?![А-Яа-я])')),
        ]

        self.address_markers = re.compile(
            r'(?i)\b(?:адрес[у|е|а]?|по\s+адресу|проживаю|прописан[а]?)\s*[:\-]?\s*'
            r'([А-Яа-яЁё][А-Яа-яЁё0-9\s,.\-/]+?(?:д\.\s*\d+[а-яА-Я/]*(?:\s*,?\s*(?:кв|к|стр|корп)\.?\s*\d+)?))',
            re.IGNORECASE
        )

        self.address_pattern = re.compile(
            r'(?i)([А-ЯЁ][а-яё]+(?:ая|ий|ое|ая)\s+)?'
            r'(?:область|обл\.|край|респ\.|республика)'
            r'[А-Яа-яЁё0-9\s,.\-/]*?'
            r'(?:д\.\s*\d+[а-яА-Я/]*(?:\s*,?\s*(?:кв|к|стр|корп)\.?\s*\d+)?)'
        )

    def _find_regex_spans(self, text: str) -> list:
        spans = []
        for pii_type, pattern in self.regex_patterns:
            for m in pattern.finditer(text):
                spans.append((m.start(), m.end(), pii_type))

        for m in self.address_markers.finditer(text):
            spans.append((m.start(), m.end(), 'ADDRESS'))

        for m in self.address_pattern.finditer(text):
            if len(m.group()) > 15:
                spans.append((m.start(), m.end(), 'ADDRESS'))

        return spans

    def _find_ner_spans(self, text: str) -> list:
        doc = Doc(text)
        doc.segment(self.segmenter)
        doc.tag_ner(self.ner_tagger)

        spans = []
        for span in doc.spans:
            if span.type in ('PER', 'LOC', 'ORG'):
                spans.append((span.start, span.stop, span.type))
        return spans

    def _merge_spans(self, spans: list, text_len: int) -> list:
        if not spans:
            return []

        spans = sorted(spans, key=lambda x: (x[0], -x[1]))
        merged = []

        for start, end, pii_type in spans:
            if merged and start < merged[-1][1]:
                prev_start, prev_end, prev_type = merged[-1]
                if end > prev_end:
                    merged[-1] = (prev_start, end, prev_type)
            else:
                merged.append((start, end, pii_type))

        return merged

    def mask(self, text: str) -> str:
        if not text or not isinstance(text, str):
            return text

        regex_spans = self._find_regex_spans(text)
        merged_regex = self._merge_spans(regex_spans, len(text))

        temp_text = text
        offset = 0
        placeholder_map = {}

        for i, (start, end, pii_type) in enumerate(sorted(merged_regex, key=lambda x: x[0])):
            placeholder = f"__PH{i}__"
            placeholder_map[placeholder] = f"[{pii_type}]"
            adj_start = start + offset
            adj_end = end + offset
            temp_text = temp_text[:adj_start] + placeholder + temp_text[adj_end:]
            offset += len(placeholder) - (end - start)

        ner_spans = self._find_ner_spans(temp_text)

        for start, end, pii_type in sorted(ner_spans, key=lambda x: x[0], reverse=True):
            chunk = temp_text[start:end]
            if not chunk.startswith("__PH"):
                temp_text = temp_text[:start] + f'[{pii_type}]' + temp_text[end:]

        for placeholder, replacement in placeholder_map.items():
            temp_text = temp_text.replace(placeholder, replacement)

        return temp_text


def process_xlsx(in_path: str, out_path: str, sheet_name: str = None, cols: list = None):
    from openpyxl import load_workbook
    from tqdm import tqdm

    cols = cols or ['A', 'B']
    masker = RuPIIMasker()

    wb = load_workbook(in_path)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    total = ws.max_row * len(cols)
    with tqdm(total=total, desc="Masking PII") as pbar:
        for r in range(1, ws.max_row + 1):
            for c in cols:
                cell = ws[f"{c}{r}"]
                if isinstance(cell.value, str) and cell.value:
                    cell.value = masker.mask(cell.value)
                pbar.update(1)

    wb.save(out_path)
    print(f"Saved: {out_path}")


def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python mask_pii_nlp.py 'text to mask'")
        print("  python mask_pii_nlp.py input.xlsx output.xlsx [sheet] [col1 col2 ...]")
        print("  echo 'text' | python mask_pii_nlp.py -")
        sys.exit(1)

    arg = sys.argv[1]

    if arg == "-":
        masker = RuPIIMasker()
        for line in sys.stdin:
            print(masker.mask(line.rstrip('\n')))
    elif arg.endswith('.xlsx'):
        if len(sys.argv) < 3:
            print("Need output.xlsx path")
            sys.exit(1)
        out_path = sys.argv[2]
        sheet = sys.argv[3] if len(sys.argv) >= 4 else None
        cols = [c.upper() for c in sys.argv[4:]] if len(sys.argv) >= 5 else ['A', 'B']
        process_xlsx(arg, out_path, sheet, cols)
    else:
        masker = RuPIIMasker()
        print(masker.mask(arg))


if __name__ == "__main__":
    main()